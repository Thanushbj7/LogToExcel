import re
import openpyxl

# Define the log file path
log_file_path = ""

# Define the Excel file path
excel_file_path = "extract-excel.xlsx"


# Regular expression for extracting date and time
datetime_pattern = r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})'  # Assuming datetime format is YYYY-MM-DD HH:MM:SS

# Regular expression for extracting a number before the word 'successful'
number_pattern = r'(\d+)\s+successful'  # Assuming the number is followed by ' successful'

# Function to extract data from the log file based on TABLE_NAME column in Excel
def extract_data_from_log(log_file_path, excel_file_path):
    
    
    with open(log_file_path, 'r') as file:
        data = file.readlines()
        first_record = re.search(datetime_pattern, data[0])
        last_record = re.findall(datetime_pattern, data[-1])
        decrypting_indexes = [i for i, line in enumerate(data) if 'Decrypting...' in line]
        decrypting_record = None
        for idx in decrypting_indexes:
            if idx + 1 < len(data):
                match = re.search(datetime_pattern, data[idx + 1])
                if match:
                    decrypting_record = match.group(0)
                    break
        last_successful = None
        for line in reversed(data):
            match = re.search(number_pattern, line)
            if match:
                last_successful = match.group(1)
                break
        return first_record.group(0), last_record[-1], decrypting_record, last_successful

# Function to update Excel file with extracted data
def update_excel_with_data(row_number,excel_file_path, first_datetime, last_datetime, decrypting_datetime, last_successful):
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb.active
    #sheet.cell(row=1, column=1).value = "First Record Date and Time"
    sheet.cell(row=row_number, column=1).value = first_datetime
    #sheet.cell(row=2, column=1).value = "Last Record Date and Time"
    sheet.cell(row=row_number, column=7).value = last_datetime
    #sheet.cell(row=3, column=1).value = "Decrypting Record Date and Time"
    sheet.cell(row=row_number, column=6).value = decrypting_datetime
    #sheet.cell(row=4, column=1).value = "Last Successful Number"
    sheet.cell(row=row_number, column=5).value = last_successful
    wb.save(excel_file_path)
    print("Data successfully updated in Excel file.")


wb = openpyxl.load_workbook(excel_file_path)
sheet = wb.active
table_names = [sheet.cell(row=i, column=3).value for i in range(2, sheet.max_row + 1)]
print(table_names)
if 'REG_MAP' in table_names:
    log_file_path = "example.log"
    first_datetime, last_datetime, decrypting_datetime, last_successful = extract_data_from_log(log_file_path, excel_file_path)
    # Update the Excel file with the extracted data
    if first_datetime and last_datetime and decrypting_datetime and last_successful:
        update_excel_with_data(2,excel_file_path, first_datetime, last_datetime, decrypting_datetime, last_successful)
if "asa" in table_names:
    log_file_path = "example1.log"
    first_datetime, last_datetime, decrypting_datetime, last_successful = extract_data_from_log(log_file_path, excel_file_path)
    if first_datetime and last_datetime and decrypting_datetime and last_successful:
        update_excel_with_data(3,excel_file_path, first_datetime, last_datetime, decrypting_datetime, last_successful)
if "start_map2" in table_names:
    log_file_path = "example2.log"
    first_datetime, last_datetime, decrypting_datetime, last_successful = extract_data_from_log(log_file_path, excel_file_path)
    if first_datetime and last_datetime and decrypting_datetime and last_successful:
        update_excel_with_data(4,excel_file_path, first_datetime, last_datetime, decrypting_datetime, last_successful)
if 'REG_MAP' not in table_names and "asa" not in table_names and "start_map2" not in table_names:
    print("No records found with TABLE_NAME as REG_MAP in the Excel file.")
    first_datetime, last_datetime, decrypting_datetime, last_successful= None, None, None, None
# Extract data from the log file based on TABLE_NAME column in Excel



