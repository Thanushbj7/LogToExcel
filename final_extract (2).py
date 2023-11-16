import re
import openpyxl
import zipfile
import os


######################################################
# Regular expression for extracting date and time
# Assuming datetime format is YYYY-MM-DD HH:MM:SS
datetime_pattern = r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})'
# Regular expression for extracting a number before the word 'successful'
# Assuming the number is followed by ' successful'
number_pattern = r'(\d+)\s+successful'
######################################################


######################################################
# Function to extract data from the log file based on TABLE_NAME column in Excel
def extract_data_from_log(log_file_path, excel_file_path):
    with open(log_file_path, 'r') as file:
        data = file.readlines()
        first_record = re.search(datetime_pattern, data[0])
        last_record = re.findall(datetime_pattern, data[-1])
        decrypting_indexes = [i for i, line in enumerate(
            data) if 'Decrypting...' in line]
        decrypting_record = None
        for idx in decrypting_indexes:
            if idx + 1 < len(data):
                match = re.search(datetime_pattern, data[idx + 1])
                if match:
                    decrypting_record = match.group(0)
                    break
        last_successful = None
        for line in reversed(data):
            match = re.search(number_pattern, line)
            if match:
                last_successful = match.group(1)
                break
        return first_record.group(0), last_record[-1], decrypting_record, last_successful
######################################################


######################################################
# Function to update Excel file with extracted data
def update_excel_with_data(table,row_number, excel_file_path, first_datetime, last_datetime, decrypting_datetime, last_successful, result):
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb.active
    for row_no in row_number:
        # sheet.cell(row=1, column=1).value = "First Record Date and Time"
        sheet.cell(row=row_no, column=1).value = first_datetime
        # sheet.cell(row=2, column=1).value = "Last Record Date and Time"
        sheet.cell(row=row_no, column=8).value = last_datetime
        # sheet.cell(row=3, column=1).value = "Decrypting Record Date and Time"
        sheet.cell(row=row_no, column=7).value = decrypting_datetime
        # sheet.cell(row=4, column=1).value = "Last Successful Number"
        sheet.cell(row=row_no, column=6).value = last_successful
        sheet.cell(row=row_no, column=2).value = result
        if table=="SFDC_W_FINANCIAL ACCOUNT" and table=="SFDC_W_FINANCIAL_ACCOUNT_TEAM":
         sheet.cell(row=row_no, column=5).value = "W"
        else:
         sheet.cell(row=row_no, column=5).value = "D"
        wb.save(excel_file_path)
        print("Data successfully updated in Excel file.")
######################################################


######################################################
def count_occurrences_in_zip_file_names(folder_path, target_string):
    # Initialize a counter for the number of occurrences
    occurrences_count = 0

    # Iterate over each file in the folder
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            # Check if the file is a zip file
            if file.endswith(".zip"):
                # Check if the target string is present in the zip file name
                if target_string in file:
                    occurrences_count += 1

    return occurrences_count
######################################################

######################################################


def extract_zip(zip_file_path, extract_folder):

    # Create the extraction folder if it doesn't exist
    if not os.path.exists(extract_folder):
        os.makedirs(extract_folder)

    # Unzip the folder
    with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
        zip_ref.extractall(extract_folder)
#########################################################

######################################################


def read_files(t_name):
    # Specify the path to the zip file
    # zip_file_path = 'zipfiles.zip'
    extract_folder = './zipfiles.zip'
    # extract_zip(zip_file_path, extract_folder)

    for root, dirs, files in os.walk(extract_folder):
        for file in files:
            # print("files in first zip folder", file)
            if file.endswith(".zip"):
                nested_extract_folder = './zipfiles/zipfiles/' + \
                    file[:-4]
                print("zip within zip, ", nested_extract_folder)
                extract_zip('./zipfiles/zipfiles/'+file, nested_extract_folder)
                for root1, dirs1, files1 in os.walk(nested_extract_folder):
                    for nested_file in files1:
                        if "sfdcRegistrationMapExtractProcess" in nested_file and t_name == "SFDC_W_TR_REGISTRATION_MAP ":
                            # print("entered first if")
                            file_path = os.path.join(root1, nested_file)
                            print([file_path, [2, 21, 30]])
                            return [file_path, [2, 21, 30]]
                        if "sfdcSbrLetterLogRelExtractProcess" in nested_file and t_name == "SBR_W_REG_LETTER_LOG_REL_SFDC":
                            file_path = os.path.join(root1, nested_file)
                            print([file_path, [3]])
                            return [file_path, [3,22,31]]
                        if "sfdcSponserNamesExtractProcess" in nested_file and t_name == "SFDC_W_SPONSOR_NAMES":
                            file_path = os.path.join(root1, nested_file)
                            return [file_path, [4,23,32]]
                        if "sfdcClientExtractProcess" in nested_file and t_name == "SFDC_W_CLIENT":
                            file_path = os.path.join(root1, nested_file)
                            return [file_path, [5,24,33]]
                        if "sfdcRegistrationExtractProcess" in nested_file and t_name == "SFDC_W_REGISTRATION":
                            file_path = os.path.join(root1, nested_file)
                            return [file_path, [6,25,34]]
                        if "oracleEblotterExtractProcess" in nested_file and t_name == "SFDC_EBLOTTER":
                            file_path = os.path.join(root1, nested_file)
                            return [file_path, [16,15]]
                        if "sfdcRegMemberExtractProcess" in nested_file and t_name == "SFDC_W_REGISTRATION_MEMBERS":
                            file_path = os.path.join(root1, nested_file)
                            return [file_path, [7,26,35]]
                        if "sfdcRegBeneficiaryExtractProcess" in nested_file and t_name == "SFDC_W_BENEFICIARY":
                            file_path = os.path.join(root1, nested_file)
                            return [file_path, [8,27,36]]
                        if "sfdcClientDisclosureExtractProcess" in nested_file and t_name == "SFDC_W_CLIENT_DISCLOSURE":
                            file_path = os.path.join(root1, nested_file)
                            return [file_path, [9,28,37]]
                        if "sfdcPortfolioReviewExtractProcess" in nested_file and t_name == "SFDC_W_PORTFOLIO_REVIEW":
                            file_path = os.path.join(root1, nested_file)
                            return [file_path, [10,29,38]]
                        if "SFDCHistoryAccountHistoryExtract" in nested_file and t_name == "SBR_ACCOUNT_HISTORY_SFDC":
                            file_path = os.path.join(root1, nested_file)
                            return [file_path, [11]]
                        if "SFDCHistoryRegClientmemberHistoryExtract" in nested_file and t_name == "SBR_REG_MEMBER_HISTORY_SFDC":
                            file_path = os.path.join(root1, nested_file)
                            return [file_path, [12]]
                        if "SFDCHistoryRegistrationHistoryExtract" in nested_file and t_name == "SBR_REGISTRATION_HISTORY_SFDC":
                            file_path = os.path.join(root1, nested_file)
                            return [file_path, [13]]
                        if "SFDCHistoryRegistrationLogExtract" in nested_file and t_name == "SBR_REG_LETTER_LOG_SFDC":
                            file_path = os.path.join(root1, nested_file)
                            return [file_path, [14]]
                        if "SFDCHistoryRegistrationLogtable_T2_Extract" in nested_file and t_name == "SBR_REG_LETTER_LOG_T2_SFDC":
                            file_path = os.path.join(root1, nested_file)
                            return [file_path, [15]]
                        if "sfdcEBlotterChecksExtractProcess" in nested_file and t_name == "SFDC_CHECKS":
                            file_path = os.path.join(root1, nested_file)
                            return [file_path, [18]]
                        if "sfdcEBlotterTradesExtractProcess" in nested_file and t_name == "SFDC_TRADES":
                            file_path = os.path.join(root1, nested_file)
                            return [file_path, [19]]
                       



######################################################


######################################################
# Define the log file path
log_file_path = ""
# Define the Excel file path
excel_file_path = "./Uzipthezip.xlsx"
wb = openpyxl.load_workbook(excel_file_path)
sheet = wb.active
table_names = [sheet.cell(
    row=i, column=3).value for i in range(2, sheet.max_row + 1)]
print(table_names)

for table in table_names:

    log_file_path, row_number = read_files(table)
    print(log_file_path)
    first_datetime, last_datetime, decrypting_datetime, last_successful = extract_data_from_log(
        log_file_path, excel_file_path)
    # print("jdiwhjdiwjdiwejdiejdiej", table)
    # Update the Excel file with the extracted data
    if first_datetime and last_datetime and decrypting_datetime and last_successful:
        # Replace 'folder_path' with the path to your folder containing zip files
        folder_path = 'zipfiles\zipfiles'
        # Replace 'ABC' with the target string you want to count
        target_string = ""
        if table in ["SFDC_W_TR_REGISTRATION_MAP ", "SBR_W_REG_LETTER_LOG_REL_SFDC", "SFDC_W_SPONSOR_NAMES"]:
            target_string = 'Copy_Tables_extract_log_files'
        if table in ["SFDC_W_CLIENT", "SFDC_W_REGISTRATION","SFDC_W_REGISTRATION_MEMBERS","SFDC_W_BENEFICIARY","SFDC_W_CLIENT_DISCLOSURE","SFDC_W_PORTFOLIO_REVIEW"]:
            target_string = 'trade_review_extract_log_files'
        if table == ["SFDC_EBLOTTER"]:
            # print("eb plotter")
            target_string = 'EblotterExtract__log_files'
        if table in ["SBR_ACCOUNT_HISTORY_SFDC","SBR_REG_MEMBER_HISTORY_SFDC","SBR_REGISTRATION_HISTORY_SFDC","SBR_REG_LETTER_LOG_SFDC","SBR_REG_LETTER_LOG_T2_SFDC"]:
            target_string = 'SFDC_History'
        if table == ["SFDC_CHECKS","SFDC_TRADES"]:
            target_string = 'sfdcEBlotter'
        

        # Count the occurrences of the target string in zip file names
        result = count_occurrences_in_zip_file_names(
            folder_path, target_string)

        print(
            f'The string "{target_string}" appears in the names of {result} zip files in the folder.')
        update_excel_with_data(table,row_number, excel_file_path, first_datetime,
                               last_datetime, decrypting_datetime, last_successful, result)
######################################################
